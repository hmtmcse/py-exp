import time
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill, Font, Side, Border, NamedStyle


def basic():
    workbook = Workbook()
    active_sheet = workbook.active
    active_sheet.title = "Basic Sheet"

    data_bank = [
        [10, 11, 12, 13],
        [14, 15, 16, 17],
        [18, 19, 20, 21],
    ]

    for data in data_bank:
        active_sheet.append(data)

    workbook.save("ignore/basic-openpyxl.xlsx")


def add_image():
    workbook = Workbook()
    active_sheet = workbook.active
    active_sheet.title = "Image Sheet"

    logo = Image("assets/logo.png")
    logo.height = 150
    logo.width = 150

    active_sheet.add_image(logo, "A3")
    workbook.save("ignore/image-openpyxl.xlsx")


def add_data_to_cell():
    workbook = Workbook()
    active_sheet = workbook.active
    active_sheet['A1'] = 87
    active_sheet['A2'] = "Name of things"
    active_sheet['A3'] = 41.80
    active_sheet['A4'] = 10

    now = time.strftime("%x")
    active_sheet['A5'] = now
    workbook.save("ignore/add_data_to_cell.xlsx")


def play_with_sheet():
    workbook = Workbook()

    active_sheet = workbook.active
    active_sheet.title = "Default"
    active_sheet['A1'] = "Default"

    workbook.create_sheet('Result')
    workbook.create_sheet('Mark')
    workbook.create_sheet('Test')

    result_sheet = workbook["Result"]
    result_sheet['A1'] = "Result"

    mark_sheet = workbook["Mark"]
    mark_sheet['A1'] = "Mark"

    test_sheet = workbook["Test"]
    test_sheet['A1'] = "Test"

    workbook.save("ignore/play_with_sheet.xlsx")


def merge_cell():
    workbook = Workbook()
    active_sheet = workbook.active

    # Way 1
    active_sheet.merge_cells('A1:B2')
    merged_cell = active_sheet["A1"]
    merged_cell.value = "Merged by Way 1"

    # Way 2
    active_sheet.merge_cells(start_row=4, start_column=1, end_row=6, end_column=5)
    merged_cell = active_sheet.cell(row=4, column=1)
    merged_cell.value = "Merged by Way 2"

    # Way 3
    active_sheet.merge_cells('A10:D12')
    active_sheet['A10'] = "Merged by Way 3"

    active_sheet['A10'].alignment = Alignment(horizontal='center', vertical='center')

    fill = PatternFill(start_color="E1E7F0", end_color="E1E7F0", fill_type="solid")
    active_sheet['A10'].fill = fill

    workbook.save("ignore/merge_cell.xlsx")


def freeze_panes():
    workbook = Workbook()
    active_sheet = workbook.active
    active_sheet['A1'] = "Name"
    active_sheet['B1'] = "Age"
    active_sheet['C1'] = "Email"
    active_sheet.freeze_panes = "A2"

    workbook.save("ignore/freeze_panes.xlsx")


def styling():
    workbook = Workbook()
    active_sheet = workbook.active

    simple_font = Font(bold=True)
    active_sheet['A1'] = "Name"
    active_sheet['A1'].font = simple_font

    mid_font = Font(bold=True, color="00FF0000", size=20)
    active_sheet['B1'] = "Age"
    active_sheet['B1'].font = mid_font

    double_border_side = Side(border_style="double")
    square_border = Border(top=double_border_side, right = double_border_side, bottom = double_border_side, left = double_border_side)
    active_sheet['C1'] = "Email"
    active_sheet['C1'].border = square_border

    # Let's create a style template for the header row
    header = NamedStyle(name="header")
    header.font = Font(bold=True)
    header.border = Border(bottom=Side(border_style="thin"))
    header.alignment = Alignment(horizontal="center", vertical="center")
    active_sheet['A5'] = "Some Header"
    active_sheet['A5'].style = header

    workbook.save("ignore/styling.xlsx")


def invoice_with_header():
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Load and add the logo image
    logo = Image("assets/logo.png")
    ws.add_image(logo, "A1")

    # Adjust the row height and column width for the logo and header
    ws.row_dimensions[1].height = 50  # Adjust based on logo size
    ws.column_dimensions["A"].width = 15

    # Add company name
    ws["B1"] = "Your Company Name"
    ws["B1"].font = Font(size=16, bold=True)
    ws["B1"].alignment = Alignment(horizontal="left", vertical="top")

    # Add company address
    ws["B2"] = "1234 Elm Street"
    ws["B2"].font = Font(size=12)
    ws["B2"].alignment = Alignment(horizontal="left")

    # Add city, state, ZIP
    ws["B3"] = "City, State ZIP"
    ws["B3"].font = Font(size=12)
    ws["B3"].alignment = Alignment(horizontal="left")

    # Add contact information
    ws["B4"] = "Phone: (123) 456-7890 | Email: info@company.com"
    ws["B4"].font = Font(size=12)
    ws["B4"].alignment = Alignment(horizontal="left")

    # Save the workbook
    wb.save("ignore/invoice_with_header.xlsx")


invoice_with_header()
styling()
freeze_panes()
merge_cell()
play_with_sheet()
add_data_to_cell()
basic()
add_image()

# https://www.blog.pythonlibrary.org/2021/08/11/styling-excel-cells-with-openpyxl-and-python/
