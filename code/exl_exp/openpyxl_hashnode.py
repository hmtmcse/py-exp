import csv
import openpyxl
from openpyxl.packaging.core import DocumentProperties
from openpyxl.styles import Side, PatternFill, Font, Alignment, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula


def load_csv(csv_file):
    # Create a new OpenPyXL Workbook
    wb = openpyxl.Workbook()

    # Select the active sheet
    ws = wb.active

    # open the `csv_file`
    with open(csv_file) as f:
        # Read the CSV data
        reader = csv.reader(f, delimiter=',')

        # Load the data in the worksheet
        for row in reader:
            ws.append(row)
    return wb


def clean_wb(wb):
    # Set up some Excel Properties
    wb.properties = DocumentProperties(
        creator="Russ",
        title="Country List",
        lastModifiedBy="OpenPyXL")

    # select the active page
    sheet = wb.active

    # Provide a good name for the sheet
    sheet.title = "Country Data"

    # iterate over every column
    for column_cells in sheet.columns:

        # iterate over evey cell in the column
        for cell in column_cells:

            # if the cell has data
            if cell.value:
                # strip spaces at the start and end
                cell.value = cell.value.strip()


def autofit(ws):

    # Iterate over the columns in the worksheet
    for column_cells in ws.columns:

        # Determine the longest item in each column
        max_length = max([len(cell.value or "") for cell in column_cells])

        # I'm adding some extra since I like it a bit roomy
        max_length = (max_length + 2) * 1.2

        # set the width of the column to the max_length
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length


def format_header_cell(cell):

    # to make things a little clearer these are the colors I'm using for styling
    blue = "000000FF"
    lightBlue = "0099CCFF"
    black = "00000000"
    white = "00FFFFFF"

    # Here we are defining the border style
    thin_border = Side(border_style="thin", color=blue)
    double_border = Side(border_style="double", color=lightBlue)

    # apply a fill to the cell
    cell.fill = PatternFill(start_color=blue, end_color=blue, fill_type="solid")

    # apply the font styles
    cell.font = Font(name="Tahoma", size=12, color=white, bold=True)

    # apply an alignment
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # apply the border
    cell.border = Border(
        top=double_border, left=thin_border, right=thin_border, bottom=double_border
    )


def style_header_row(wb):
    # Select the active page
    ws = wb.active

    # The column header is in the first row
    column_header_row = 1

    # Select the column headers
    for cell in ws[column_header_row]:
        format_header_cell(cell)

    # Use the user defined function to set the column width
    autofit(ws)

    # Adjust Column 'H' due to it's width
    ws.column_dimensions["H"].width = ws.column_dimensions["H"].width * 0.6


def country_column(wb):
    # Select the active page
    ws = wb.active

    # Countries are in the first column
    country_column = "A"

    # To make things a little clearer these are the colors I'm using for styling
    babyBlue = "00CCFFFF"

    # Here we are defining the border style
    thick_border = Side(border_style="thick")

    # Iterating over the Country column skipping the header row
    for cell in ws[country_column][1:]:
        # Formatting each cell
        cell.fill = PatternFill(
            start_color=babyBlue, end_color=babyBlue, fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="right")
        cell.border = Border(right=thick_border)


def format_numbers(wb):
    # Select the active page
    ws = wb.active

    # 2 types of numeric columns: no decimals and 2 decimals
    number_col_0_dec = ['C', 'D', 'I']

    # Iterate over the number columns
    for col in ws['C:T']:

        # Iterate over the cells in that column, skip the header row
        for cell in col[1:]:

            # Verify the cell has a value and is of instance `str`
            if cell.value and isinstance(cell.value, str):
                # replace the comma with a period
                cell.value = float(cell.value.replace(',', '.'))

            # apply format
            cell.number_format = "#,##0" if col in number_col_0_dec else "#,##0.00"


def ranking_sheet(wb):
    # add a summary sheet
    rankings_sheet = wb.create_sheet(title="Rankings")

    # Header Row
    rankings_sheet['A1'] = "Ranking"
    format_header_cell(rankings_sheet['A1'])
    rankings_sheet.column_dimensions['A'].width = 25

    format_header_cell(rankings_sheet['B1'])
    rankings_sheet['B1'] = "Country"
    rankings_sheet.column_dimensions['B'].width = 20

    format_header_cell(rankings_sheet['C1'])
    rankings_sheet['C1'] = "Values"
    rankings_sheet.column_dimensions['C'].width = 15

    # Largest Population
    rankings_sheet['A2'] = "Largest population"
    rankings_sheet[
        'B2'] = "=INDEX('Country Data'!A2:A229, MATCH(MAX('Country Data'!C2:C228), 'Country Data'!C2:C228, 0))"
    rankings_sheet['C2'] = "=MAX('Country Data'!C2:C228)"
    rankings_sheet['C2'].number_format = "#,###"

    # Largest Population
    rankings_sheet['A3'] = "Smallest population"
    rankings_sheet[
        'B3'] = "=INDEX('Country Data'!A2:A229, MATCH(MIN('Country Data'!C2:C228), 'Country Data'!C2:C228, 0))"
    rankings_sheet['C3'] = "=MIN('Country Data'!C2:C228)"
    rankings_sheet['C3'].number_format = "#,###"

    # Largest Area
    rankings_sheet['A4'] = "Largest area"
    rankings_sheet[
        'B4'] = "=INDEX('Country Data'!A2:A229, MATCH(MAX('Country Data'!D2:D228), 'Country Data'!D2:D228, 0))"
    rankings_sheet['C4'] = "=MAX('Country Data'!D2:D228)"
    rankings_sheet['C4'].number_format = "#,###"

    # Smallest Area
    rankings_sheet['A5'] = "Smallest area"
    rankings_sheet[
        'B5'] = "=INDEX('Country Data'!A2:A229, MATCH(MIN('Country Data'!D2:D228), 'Country Data'!D2:D228, 0))"
    rankings_sheet['C5'] = "=MIN('Country Data'!D2:D228)"
    rankings_sheet['C5'].number_format = "#,###"


def summary_sheet(wb):
    # add a summary sheet
    summary_sheet = wb.create_sheet(title="Summary")

    # Region header
    summary_sheet["A1"].value = "Region"
    format_header_cell(summary_sheet["A1"])

    # per OpenPyXL's documention, any formula that wasn't in the initial specification
    # must be prefixed with `_xlfn.`
    summary_sheet["A2"] = ArrayFormula(
        "A2:A12", "=_xlfn.UNIQUE(_xlfn.SORT('Country Data'!B2:B228))"
    )

    summary_sheet.column_dimensions["A"].width = 22

    # Count of Country
    summary_sheet["B1"].value = "# of Countries"
    format_header_cell(summary_sheet["B1"])

    for col in summary_sheet["B2:B12"]:
        for cell in col:
            cell.value = f"=COUNTIF('Country Data'!$B$2:$B$228, A{cell.row})"
            cell.number_format = "#,###"

    summary_sheet.column_dimensions["B"].width = 15

    # Population
    summary_sheet["C1"].value = "Population"
    format_header_cell(summary_sheet["C1"])

    for col in summary_sheet["C2:C12"]:
        for cell in col:
            cell.value = f"=SUMIF('Country Data'!$B$2:$B$228, A{cell.row}, 'Country Data'!$C$2:$C$228)"
            cell.number_format = "#,###"

    summary_sheet.column_dimensions["C"].width = 15

    # Area
    summary_sheet["D1"].value = "Area"
    format_header_cell(summary_sheet["D1"])

    for col in summary_sheet["D2:D12"]:
        for cell in col:
            cell.value = f"=SUMIF('Country Data'!$B$2:$B$228, A{cell.row}, 'Country Data'!$D$2:$CD$228)"
            cell.number_format = "#,###"

    summary_sheet.column_dimensions["D"].width = 15


def main(csv_file, xslx_file):
    workbook = load_csv(csv_file)

    # calling each function in turn
    clean_wb(workbook)
    style_header_row(workbook)
    # country_column(workbook)
    # format_numbers(workbook)
    # ranking_sheet(workbook)
    # summary_sheet(workbook)

    # save the OpenPyXL Workbook
    workbook.save(xslx_file)


if __name__ == "__main__":
    main("assets/countries-of-the-world.csv", 'ignore/countries-report.xlsx')
