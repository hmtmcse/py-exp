from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill


def print_invoice1():

    workbook = Workbook()
    active_sheet = workbook.active
    active_sheet.title = "Invoice"

    active_sheet.merge_cells('A1:A3')
    # active_sheet.column_dimensions['A'].width = 11
    logo = Image("assets/logo.png")
    logo.height = 75
    logo.width = 75
    active_sheet.add_image(logo, "A1")
    active_sheet["A1"].alignment = Alignment(vertical="center")

    active_sheet.merge_cells('B2:J2')
    active_sheet["B2"].alignment = Alignment(horizontal="center", vertical="center")
    active_sheet['B2'] = "Bangla Fighter School & College, Rangpur"
    active_sheet['B2'].font = Font(bold=True, size=16, name='Arial')
    active_sheet.row_dimensions[2].height = 30

    active_sheet.merge_cells('B3:J3')
    active_sheet["B3"].alignment = Alignment(horizontal="center", vertical="center")
    active_sheet['B3'] = "Rangpur Sadar"
    active_sheet['B3'].font = Font(bold=True, size=12, name='Arial')
    active_sheet.row_dimensions[3].height = 22

    thick_side = Side(border_style="thick", color="000000")
    border = Border(top=thick_side)
    for row in active_sheet['A4:J4']:
        for cell in row:
            cell.border = border

    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for row in active_sheet['A1:J3']:
        for cell in row:
            cell.fill = white_fill

    workbook.save("ignore/print_invoice1.xlsx")


print_invoice1()
